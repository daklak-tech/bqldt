#!/usr/bin/env python3
from __future__ import annotations

import argparse
import json
import re
import sys
import tempfile
import urllib.request
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


DEFAULT_SHEET_URL = "https://docs.google.com/spreadsheets/d/1vpDSRjzNjJL3gNsAOdIxFh0Aad16yoD4/edit?gid=631738297#gid=631738297"
MONTH_RE = re.compile(r"th[aá]ng\s*(\d{1,2})", re.I)
PARTIAL_RE = re.compile(r"\(\s*(\d{1,2})\s*/\s*(\d{1,2})(?:\s*/\s*(\d{2,4}))?\s*\)")
COMPACT_PARTIAL_RE = re.compile(r"\(\s*(\d{3,4})\s*\)")


def norm(value: Any) -> str:
    import unicodedata

    text = str(value or "").strip().lower()
    text = "".join(ch for ch in unicodedata.normalize("NFD", text) if unicodedata.category(ch) != "Mn")
    return text.replace("đ", "d")


def num(value: Any) -> int | None:
    if value is None or value == "":
        return None
    if isinstance(value, (int, float)):
        return int(round(value))
    cleaned = re.sub(r"[^\d.-]", "", str(value))
    if not cleaned:
        return None
    return int(round(float(cleaned)))


def sheet_export_url(url: str) -> str:
    m = re.search(r"/spreadsheets/d/([^/]+)", url)
    if not m:
        return url
    return f"https://docs.google.com/spreadsheets/d/{m.group(1)}/export?format=xlsx"


def download_sheet(url: str) -> Path:
    target = Path(tempfile.gettempdir()) / "bqldt-source.xlsx"
    request = urllib.request.Request(sheet_export_url(url), headers={"User-Agent": "Mozilla/5.0"})
    with urllib.request.urlopen(request, timeout=60) as response:
        target.write_bytes(response.read())
    return target


def extract_payload(html: str) -> dict[str, Any]:
    m = re.search(r'<script id="dashboard-data" type="application/json">(.*?)</script>', html, re.S)
    if not m:
        raise SystemExit("Không tìm thấy khối dashboard-data trong index.html")
    return json.loads(m.group(1))


def replace_payload(html: str, data: dict[str, Any]) -> str:
    payload = json.dumps(data, ensure_ascii=False, separators=(",", ":")).replace("</", "<\\/")
    return re.sub(
        r'(<script id="dashboard-data" type="application/json">).*?(</script>)',
        lambda m: m.group(1) + payload + m.group(2),
        html,
        flags=re.S,
    )


def monument_from_sheet(title: str) -> tuple[str, str, str] | None:
    n = norm(title)
    if "ganh" in n and "da" in n:
        return ("GÀNH ĐÁ ĐĨA", "gdd25", "gdd26")
    if "bai" in n and ("mon" in n or "mui" in n):
        return ("BÃI MÔN - MŨI ĐẠI LÃNH", "bm25", "bm26")
    if "thap" in n and "nhan" in n:
        return ("THÁP NHẠN", "", "tn26")
    return None


def read_month_rows(ws) -> list[dict[str, Any]]:
    revenue_cols: list[int] = []
    visitor_cols: list[int] = []
    for header in ws.iter_rows(min_row=1, max_row=min(ws.max_row, 8), values_only=True):
        normalized = [norm(cell) for cell in header]
        if not any("doanh thu" in cell or "luot khach" in cell for cell in normalized):
            continue
        for idx, text in enumerate(normalized):
            if "doanh thu" in text:
                revenue_cols.append(idx)
            if "luot khach" in text:
                visitor_cols.append(idx)
        if revenue_cols or visitor_cols:
            break

    revenue25_col = revenue_cols[0] if len(revenue_cols) >= 1 else 1
    revenue26_col = revenue_cols[1] if len(revenue_cols) >= 2 else 3
    visitors25_col = visitor_cols[0] if len(visitor_cols) >= 1 else 2
    visitors26_col = visitor_cols[1] if len(visitor_cols) >= 2 else 4

    def get(row: tuple[Any, ...], col: int | None) -> int | None:
        if col is None or col >= len(row):
            return None
        return num(row[col])

    rows: list[dict[str, Any]] = []
    for row in ws.iter_rows(values_only=True):
        label = None
        m = None
        for cell in row:
            m = MONTH_RE.search(str(cell or ""))
            if m:
                label = cell
                break
        if not m:
            continue
        rows.append(
            {
                "month_index": int(m.group(1)) - 1,
                "label": str(label).strip(),
                "revenue25": get(row, revenue25_col),
                "visitors25": get(row, visitors25_col),
                "revenue26": get(row, revenue26_col),
                "visitors26": get(row, visitors26_col),
            }
        )
    return rows


def partial_key(label: str) -> tuple[int, int, int] | None:
    m = PARTIAL_RE.search(label)
    if m:
        day, month, year = int(m.group(1)), int(m.group(2)), int(m.group(3) or 2026)
        if year < 100:
            year += 2000
        return (year, month, day)

    compact = COMPACT_PARTIAL_RE.search(label)
    month_match = MONTH_RE.search(label)
    if not compact or not month_match:
        return None

    digits = compact.group(1)
    month = int(month_match.group(1))
    day: int | None = None

    # Một số ô nguồn gõ nhanh dạng "Thang 7(287)" thay vì "Tháng 7 (28/7)".
    # Ưu tiên tháng đã ghi ngoài nhãn, rồi suy ra phần còn lại là ngày.
    month_text = str(month)
    if digits.endswith(month_text):
        possible_day = digits[: -len(month_text)]
        if possible_day.isdigit():
            day = int(possible_day)
    if day is None and len(digits) == 4 and len(month_text) == 1 and digits.endswith(month_text * 2):
        possible_day = digits[:-2]
        if possible_day.isdigit():
            day = int(possible_day)

    if day is None or not (1 <= day <= 31 and 1 <= month <= 12):
        return None
    year = 2026
    return (year, month, day)


def clean_month_label(label: str) -> str:
    text = re.sub(r"\s+", " ", label).strip()
    key = partial_key(text)
    month_match = MONTH_RE.search(text)
    if key and month_match:
        _, month, day = key
        return f"Tháng {int(month_match.group(1))} ({day}/{month})"
    text = re.sub(r"\(\s*(\d{1,2})\s*/\s*(\d{1,2})(?:\s*/\s*(\d{2,4}))?\s*\)", lambda m: f"({m.group(1)}/{m.group(2)}{('/' + m.group(3)) if m.group(3) else ''})", text)
    text = re.sub(r"^th[aá]ng", "Tháng", text, flags=re.I)
    return re.sub(r"(Tháng\s*\d{1,2})\s*\(", r"\1 (", text, flags=re.I)


def vnd(value: int) -> str:
    return f"{value:,}".replace(",", ".")


def date_from_partial_label(label: str | None) -> str | None:
    if not label:
        return None
    key = partial_key(label)
    if not key:
        return None
    year, month, day = key
    return f"{day}/{month}/{year}"


def refresh_monument_result(
    data: dict[str, Any],
    monument_name: str,
    revenue_total_26: int,
    visitor_total_26: int,
    partial_label: str | None,
) -> None:
    date_text = date_from_partial_label(partial_label)
    if not date_text:
        return
    for monument in data.get("monuments", []):
        if monument.get("name") != monument_name:
            continue
        current = monument.get("result", "")
        visitors = data.get("analytics", {}).get(monument_name, {}).get("visitors") or visitor_total_26
        if monument_name == "GÀNH ĐÁ ĐĨA":
            sentence = (
                f"Theo bảng số liệu cập nhật đến ngày {date_text}, doanh thu đạt {vnd(revenue_total_26)} đồng; "
                f"lượt khách năm 2026 ghi nhận {vnd(int(visitors))} lượt."
            )
            pattern = r"Theo bảng số liệu cập nhật đến ngày \d{1,2}/\d{1,2}/2026, doanh thu đạt [0-9.]+ đồng; lượt khách năm 2026 ghi nhận [0-9.]+ lượt\."
        elif monument_name == "BÃI MÔN - MŨI ĐẠI LÃNH":
            sentence = (
                f"Theo bảng số liệu cập nhật đến ngày {date_text}, doanh thu đạt {vnd(revenue_total_26)} đồng; "
                f"lượt khách năm 2026 ghi nhận {vnd(int(visitors))} lượt."
            )
            pattern = r"Theo bảng số liệu cập nhật đến ngày \d{1,2}/\d{1,2}/2026, doanh thu đạt [0-9.]+ đồng; lượt khách năm 2026 ghi nhận [0-9.]+ lượt\."
        else:
            return
        if re.search(pattern, current):
            monument["result"] = re.sub(pattern, sentence, current)
        else:
            monument["result"] = (current.rstrip() + " " + sentence).strip()
        return


def refresh_data(data: dict[str, Any], xlsx_path: Path) -> dict[str, Any]:
    wb = load_workbook(xlsx_path, data_only=True)
    monthly = data["monthly"]
    partial_labels: dict[int, str] = {}
    monument_partial_labels: dict[str, tuple[tuple[int, int, int], str]] = {}
    latest_partial_label: tuple[tuple[int, int, int], str] | None = None
    updated_any_sheet = False
    seen_monuments: set[str] = set()

    for ws in wb.worksheets:
        mapping = monument_from_sheet(ws.title)
        if not mapping:
            continue
        monument, field25, field26 = mapping
        rows = read_month_rows(ws)
        if monument in seen_monuments:
            sheet_revenue25 = sum(item["revenue25"] or 0 for item in rows)
            sheet_revenue26 = sum(item["revenue26"] or 0 for item in rows)
            if sheet_revenue25 == 0 and sheet_revenue26 > 0:
                monument, field25, field26 = ("THÁP NHẠN", "", "tn26")
            else:
                print(f"Bỏ qua sheet trùng '{ws.title}' vì đã cập nhật {monument}.")
                continue

        v25 = [None] * 12
        v26 = [None] * 12
        revenue_total_26 = 0
        visitor_total_26 = 0
        has_revenue_data = False

        for item in rows:
            i = item["month_index"]
            if not 0 <= i < len(monthly):
                continue
            if item["revenue25"] is not None or item["revenue26"] is not None:
                has_revenue_data = True
            if field25 and item["revenue25"] is not None:
                monthly[i][field25] = item["revenue25"]
            if item["revenue26"] is not None:
                monthly[i][field26] = item["revenue26"]
            if item["revenue26"] is not None:
                revenue_total_26 += item["revenue26"]
            if item["visitors26"] is not None:
                visitor_total_26 += item["visitors26"]
            v25[i] = item["visitors25"]
            v26[i] = item["visitors26"]

            key = partial_key(item["label"])
            if key and (item["revenue26"] is not None or item["visitors26"] is not None):
                clean_label = clean_month_label(item["label"])
                old = partial_labels.get(i)
                if old is None or (partial_key(old) or (0, 0, 0)) < key:
                    partial_labels[i] = clean_label
                old_monument = monument_partial_labels.get(monument)
                if old_monument is None or old_monument[0] < key:
                    monument_partial_labels[monument] = (key, clean_label)
                if latest_partial_label is None or latest_partial_label[0] < key:
                    latest_partial_label = (key, clean_label)

        if not has_revenue_data:
            print(f"Bỏ qua sheet '{ws.title}' vì không đọc được dữ liệu doanh thu hợp lệ; giữ nguyên số liệu cũ.")
            continue

        updated_any_sheet = True
        seen_monuments.add(monument)
        if monument in data.get("analytics", {}):
            data["analytics"][monument]["revenue"] = revenue_total_26
            if visitor_total_26:
                data["analytics"][monument]["visitors"] = visitor_total_26
            refresh_monument_result(
                data,
                monument,
                revenue_total_26,
                visitor_total_26,
                monument_partial_labels.get(monument, ((0, 0, 0), None))[1],
            )

        if any(x is not None for x in v25 + v26):
            data.setdefault("visitorMonthly", {})[monument] = {
                "series": [
                    {"label": "2025", "values": v25},
                    {"label": "2026", "values": v26},
                ],
                "source": "Google Sheet cập nhật tự động",
            }

    for i, row in enumerate(monthly):
        if i in partial_labels:
            row["month"] = partial_labels[i]

    if updated_any_sheet:
        partial_notes = []
        for monument, (_, label) in monument_partial_labels.items():
            date_text = date_from_partial_label(label)
            if date_text:
                partial_notes.append(f"{monument.title()} đến {date_text}")
        for source in data.get("sources", []):
            if source.get("name") == "Bao cao 6 thang.md":
                source["note"] = "Kết quả thực hiện đến tháng hiện tại; kế hoạch thu 2026 sử dụng mốc 19,8 tỷ đồng đang thể hiện trên dashboard"
            source_name_note = f"{source.get('name', '')} {source.get('note', '')}"
            if "Nguồn Drive" in source_name_note or "Google Sheet" in source_name_note:
                source["name"] = "Google Sheet doanh thu, lượt khách 2025–2026"
                date_text = date_from_partial_label(latest_partial_label[1]) if latest_partial_label else None
                if partial_notes:
                    source["note"] = (
                        "Cập nhật tự động hằng ngày; số liệu tháng đang phát sinh: "
                        + ", ".join(partial_notes)
                        + "; Tháp Nhạn chưa có phát sinh doanh thu tháng 7 theo file nguồn."
                    )
                elif date_text:
                    source["note"] = f"Cập nhật tự động hằng ngày; số liệu tháng đang phát sinh cập nhật đến {date_text} theo file nguồn."
                else:
                    source["note"] = "Cập nhật tự động hằng ngày; bao gồm cả số liệu tháng đang phát sinh theo file nguồn."

    return data


def main() -> int:
    parser = argparse.ArgumentParser(description="Cập nhật dashboard bqldt từ Google Sheet/Excel.")
    parser.add_argument("--index", default="index.html", help="Đường dẫn index.html cần cập nhật")
    parser.add_argument("--sheet-url", default=DEFAULT_SHEET_URL, help="Google Sheet URL")
    parser.add_argument("--xlsx", help="Dùng file .xlsx cục bộ để kiểm tra")
    args = parser.parse_args()

    index_path = Path(args.index)
    html = index_path.read_text(encoding="utf-8")
    data = extract_payload(html)
    xlsx_path = Path(args.xlsx) if args.xlsx else download_sheet(args.sheet_url)
    data = refresh_data(data, xlsx_path)
    index_path.write_text(replace_payload(html, data), encoding="utf-8")
    return 0


if __name__ == "__main__":
    sys.exit(main())
